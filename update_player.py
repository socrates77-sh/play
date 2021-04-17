# history:
# 2021/3/12   v1.0  initial
# 2021/4/17   v1.1  solve GK problem


import numpy as np
import pandas as pd
import os
import msvcrt
import shutil
import datetime
import openpyxl
from openpyxl.styles import colors, Font, Color, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

VERSION = '1.1'

FIRST_TEAM_NUMBER = 22

bak_dir = './bak'
csv_file = './a.csv'
excel_file = './exeter.xlsx'
sheet_present = 'present'
sheet_leave = 'leave'


def print_version(version):
    print('=' * 70)
    print('%s version: %s' % (os.path.basename(__file__), VERSION))
    print('=' * 70)


def wait_any_key():
    print('press any key to exit...')
    msvcrt.getch()


def month_inc(txt):
    yr = int(txt[:2])
    mo = int(txt[2:])

    if mo == 12:
        mo = 1
        yr = yr+1
    else:
        mo += 1
    return '%02d%02d' % (yr, mo)


def backup_file(month_txt):
    bak_file_name = 'exeter_%s.xlsx' % month_txt
    print('%s -> %s' % (excel_file, bak_file_name))
    shutil.copy(excel_file, os.path.join(bak_dir, bak_file_name))


def find_leave(df_present, df_leave, df_newest):
    present_player = set(df_present.index.values)
    newest_player = set(df_newest.index.values)
    latest_leave_player = present_player-newest_player
    df = df_present.loc[list(latest_leave_player)]
    df = pd.concat([df_leave, df], sort=False)
    df1 = df[['Position', 'Potential']]
    df2 = df.drop(columns=['Position', 'Potential'])
    df2 = df2.sort_index(axis=1, ascending=False)
    df = pd.concat([df1, df2.reindex(df1.index)], axis=1)
    return df


def get_new_present(df_present, df_newest, month_txt):
    valid_col = ['Position', 'Potential Ability', 'Current Ability']
    df1 = df_newest[valid_col]
    df2 = df_present.drop(columns=['Position', 'Potential'])
    df1.columns = ['Position', 'Potential', month_inc(month_txt)]
    df = pd.concat([df1, df2.reindex(df1.index)], axis=1)
    df = df.sort_values(
        by=[month_inc(month_txt), 'Potential'], ascending=[False, False])
    return df


def write_xlxs(xlsx, df_present_update, df_leave_update):
    df = df_present_update
    df.to_excel(xlsx, sheet_name=sheet_present)
    df = df_leave_update
    df.to_excel(xlsx, sheet_name=sheet_leave)
    xlsx.close()


def set_width_and_fill(ws, nrows, ncols):
    fill_title1 = PatternFill(fill_type="solid", fgColor="00BFFF")
    fill_title2 = PatternFill(fill_type="solid", fgColor="FFA500")
    fill_name = PatternFill(fill_type="solid", fgColor="FFC0CB")
    fill_position = PatternFill(fill_type="solid", fgColor="90EE90")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.cell(1, 1).fill = fill_title1
    ws.cell(1, 2).fill = fill_title1
    ws.cell(1, 3).fill = fill_title1

    for i in range(3, ncols):
        ws.column_dimensions[get_column_letter(i+1)].width = 5
        ws.cell(1, i+1).fill = fill_title2

    for i in range(1, nrows):
        ws.cell(i+1, 1).fill = fill_name
        ws.cell(i+1, 2).fill = fill_position
        ws.cell(i+1, 3).fill = fill_position


def set_general_format(ws, nrows, ncols):
    font = Font(name='Arial', size=10, color=colors.BLACK)
    hd_font = Font(name='Arial', size=10, color=colors.BLACK, bold=True)
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    alignment = Alignment(horizontal="center", vertical="center")

    for i in range(nrows):
        for j in range(ncols):
            cell = ws.cell(i+1, j+1)
            cell.number_format = '0'
            cell.border = border
            cell.alignment = alignment
            if(cell.font.b):
                cell.font = hd_font
            else:
                cell.font = font


def set_first_team(ws, nrows, ncols):
    font_first_team = Font(name='Arial', size=10, color=colors.RED, bold=True)
    # n_gk_no_larger = 0
    n_found_gk = 0
    n_first = 0

    for i in range(1, nrows):
        pos = ws.cell(i+1, 2).value
        # val = ws.cell(i+1, 4).value
        if n_first >= FIRST_TEAM_NUMBER:
            break
        else:
            if ('GK' in pos):
                if n_found_gk < 2:
                    ws.cell(i+1, 1).font = font_first_team
                    n_found_gk += 1
                    n_first += 1
                else:
                    continue
            else:
                ws.cell(i+1, 1).font = font_first_team
                n_first += 1

        # if ('GK' in pos) and n_found_gk < 2 and n_first < FIRST_TEAM_NUMBER:
        #     ws.cell(i+1, 1).font = font_first_team
        #     n_found_gk += 1
        #     if i > FIRST_TEAM_NUMBER:
        #         n_gk_no_larger += 1

    # for i in range(1, FIRST_TEAM_NUMBER-n_gk_no_larger+1):
    #     ws.cell(i+1, 1).font = font_first_team


def set_value_color(ws, nrows, ncols):
    font_up = Font(name='Arial', size=10, color=colors.RED)
    font_down = Font(name='Arial', size=10, color=colors.GREEN)

    for i in range(1, nrows):
        for j in range(0, ncols-4):
            val = ws.cell(i+1, ncols-j-1).value
            val_next = ws.cell(i+1, ncols-j).value
            if val is None or val_next is None:
                continue
            if val > val_next:
                ws.cell(i+1, ncols-j-1).font = font_up
            if val < val_next:
                ws.cell(i+1, ncols-j-1).font = font_down


def format_xls(xlx_file, sheet_name):
    workbook = openpyxl.load_workbook(excel_file)
    ws = workbook[sheet_name]

    nrows = ws.max_row
    ncols = ws.max_column

    set_width_and_fill(ws, nrows, ncols)
    set_general_format(ws, nrows, ncols)
    if sheet_name == sheet_present:
        set_first_team(ws, nrows, ncols)
    set_value_color(ws, nrows, ncols)

    workbook.save(filename=excel_file)


def main():
    print_version(VERSION)

    df_newest = pd.read_csv(csv_file, encoding='gbk', index_col=0)
    df_present = pd.read_excel(
        excel_file, sheet_name=sheet_present, encoding='gbk', index_col=0)
    df_leave = pd.read_excel(
        excel_file, sheet_name=sheet_leave, encoding='gbk', index_col=0)
    month_txt = df_present.columns[2]
    backup_file(month_txt)
    xlsx = pd.ExcelWriter(excel_file)

    df_present_update = get_new_present(df_present, df_newest, month_txt)
    df_leave_update = find_leave(df_present, df_leave, df_newest)
    write_xlxs(xlsx, df_present_update, df_leave_update)

    format_xls(excel_file, sheet_present)
    format_xls(excel_file, sheet_leave)

    wait_any_key()


if __name__ == '__main__':
    main()
