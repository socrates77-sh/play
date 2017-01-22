__author__ = 'socrates'

import openpyxl


wb = openpyxl.load_workbook(r'G:\temp\a\数字组_员工周报_李秀峰.xlsx')
# print(wb.sheetnames)

# ws = wb.worksheets[0]
# print(ws.title)

c=ws['E2']
print(c.value)
print(c.number_format)

